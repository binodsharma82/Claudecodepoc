"""
IQ2020 POC — Sample Data Model Generator
Entities: MasterCustomer (HCP/HCO), Products, Territory, SalesRep,
          Speaker, Calls, Prescriptions (NRx/TRx), SalesTarget
KPIs:     Territory Performance, HCP 360 View

BU design (exclusive — no specialty shared across BUs):
  AIR      → NEUROLOGY, ONCOLOGY          → prescribes NEXOLID, VERITONEX
  SBU      → PSYCHIATRY, INTERNAL MEDICINE, GERIATRICS → prescribes CLAROZEPT, DEPTRAZOL
  Vaccines → FAMILY PRACTICE, CARDIOLOGY  → no pharma prescriptions (own endpoint)

Each territory gets exactly: 3 AIR HCPs + 5 SBU HCPs + 2 Vaccines HCPs = 100 HCPs total
"""

import os
import random
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

np.random.seed(42)
random.seed(42)

os.makedirs("sourcedata", exist_ok=True)

# ──────────────────────────────────────────────
# 1. DIM_TERRITORY
# ──────────────────────────────────────────────
TERRITORIES = [
    ("T001", "Northeast-1",    "D001", "Northeast District",     "R001", "Northeast Region",   "A001", "East Area",    "USA"),
    ("T002", "Northeast-2",    "D001", "Northeast District",     "R001", "Northeast Region",   "A001", "East Area",    "USA"),
    ("T003", "Southeast-1",    "D002", "Southeast District",     "R001", "Northeast Region",   "A001", "East Area",    "USA"),
    ("T004", "Mid-Atlantic-1", "D003", "Mid-Atlantic District",  "R002", "Mid-Atlantic Region","A001", "East Area",    "USA"),
    ("T005", "Midwest-1",      "D004", "Midwest-North District", "R003", "Central Region",     "A002", "Central Area", "USA"),
    ("T006", "Midwest-2",      "D004", "Midwest-North District", "R003", "Central Region",     "A002", "Central Area", "USA"),
    ("T007", "South-Central-1","D005", "South-Central District", "R003", "Central Region",     "A002", "Central Area", "USA"),
    ("T008", "Southwest-1",    "D006", "Southwest District",     "R004", "West Region",        "A003", "West Area",    "USA"),
    ("T009", "Pacific-1",      "D007", "Pacific District",       "R004", "West Region",        "A003", "West Area",    "USA"),
    ("T010", "Pacific-2",      "D007", "Pacific District",       "R004", "West Region",        "A003", "West Area",    "USA"),
]
df_territory = pd.DataFrame(TERRITORIES, columns=[
    "TerritoryID", "TerritoryName", "DistrictID", "DistrictName",
    "RegionID", "RegionName", "AreaID", "AreaName", "Nation",
])

# ──────────────────────────────────────────────
# 2. DIM_PRODUCTS
# ──────────────────────────────────────────────
PRODUCTS = [
    ("P001", "NEXOLID",   "NEXOLID 10MG",   "Atypical Antipsychotic", "CNS-Premium",     "PharmaCore Inc", "CNS/Neurology",  "Branded", "Y"),
    ("P002", "NEXOLID",   "NEXOLID 25MG",   "Atypical Antipsychotic", "CNS-Premium",     "PharmaCore Inc", "CNS/Neurology",  "Branded", "Y"),
    ("P003", "VERITONEX", "VERITONEX 5MG",  "Atypical Antipsychotic", "CNS-Premium",     "PharmaCore Inc", "CNS/Neurology",  "Branded", "Y"),
    ("P004", "VERITONEX", "VERITONEX 15MG", "Atypical Antipsychotic", "CNS-Premium",     "PharmaCore Inc", "CNS/Neurology",  "Branded", "Y"),
    ("P005", "CLAROZEPT", "CLAROZEPT 20MG", "Atypical Antipsychotic", "CNS-Established", "PharmaCore Inc", "CNS/Psychiatry", "Branded", "Y"),
    ("P006", "CLAROZEPT", "CLAROZEPT 40MG", "Atypical Antipsychotic", "CNS-Established", "PharmaCore Inc", "CNS/Psychiatry", "Branded", "Y"),
    ("P007", "DEPTRAZOL", "DEPTRAZOL 50MG", "Antidepressant",         "Psychiatry",      "PharmaCore Inc", "CNS/Psychiatry", "Branded", "Y"),
    ("P008", "DEPTRAZOL", "DEPTRAZOL 100MG","Antidepressant",         "Psychiatry",      "PharmaCore Inc", "CNS/Psychiatry", "Branded", "Y"),
]
df_products = pd.DataFrame(PRODUCTS, columns=[
    "ProductID", "ProductFamily", "ProductName", "TherapeuticClass",
    "ProductGroup", "Manufacturer", "TherapeuticArea", "SalesCategory", "ActiveFlag",
])

# ──────────────────────────────────────────────
# BU / Specialty / Product mapping (exclusive)
# ──────────────────────────────────────────────
BU_SPECIALTIES = {
    "AIR":      ["NEUROLOGY", "ONCOLOGY"],
    "SBU":      ["PSYCHIATRY", "INTERNAL MEDICINE", "GERIATRICS"],
    "Vaccines": ["FAMILY PRACTICE", "CARDIOLOGY"],
}
SPEC_TO_BU = {s: bu for bu, specs in BU_SPECIALTIES.items() for s in specs}

# Products each BU's HCPs actually prescribe (one representative SKU per family)
BU_PROD_IDS = {
    "AIR":      ["P001", "P003"],   # NEXOLID, VERITONEX
    "SBU":      ["P005", "P007"],   # CLAROZEPT, DEPTRAZOL
    "Vaccines": [],                  # no pharma prescriptions
}

# HCPs to generate per territory per BU
BU_HCP_QUOTA = {"AIR": 3, "SBU": 5, "Vaccines": 2}

# ──────────────────────────────────────────────
# 3. DIM_MASTER_CUSTOMER (HCP — guaranteed BU coverage per territory)
# ──────────────────────────────────────────────
FIRST_NAMES = ["James","Mary","Robert","Patricia","John","Jennifer","Michael","Linda",
               "William","Barbara","David","Susan","Richard","Jessica","Joseph","Sarah",
               "Thomas","Karen","Charles","Lisa","Christopher","Nancy","Daniel","Betty",
               "Matthew","Margaret","Anthony","Sandra","Mark","Ashley"]
LAST_NAMES  = ["Smith","Johnson","Williams","Brown","Jones","Garcia","Miller","Davis",
               "Wilson","Taylor","Anderson","Thomas","Jackson","White","Harris","Martin",
               "Thompson","Martinez","Robinson","Clark","Rodriguez","Lewis","Lee","Walker",
               "Hall","Allen","Young","Hernandez","King","Wright"]
CREDENTIALS  = ["MD","DO","NP","PA","PharmD"]
CRED_WEIGHTS = [50, 15, 15, 15, 5]
CITIES_STATES = [
    ("New York","NY"),("Los Angeles","CA"),("Chicago","IL"),("Houston","TX"),
    ("Phoenix","AZ"),("Philadelphia","PA"),("San Antonio","TX"),("San Diego","CA"),
    ("Dallas","TX"),("San Jose","CA"),("Austin","TX"),("Jacksonville","FL"),
    ("Fort Worth","TX"),("Columbus","OH"),("Charlotte","NC"),("Indianapolis","IN"),
    ("San Francisco","CA"),("Seattle","WA"),("Denver","CO"),("Nashville","TN"),
    ("Oklahoma City","OK"),("El Paso","TX"),("Las Vegas","NV"),("Louisville","KY"),
    ("Baltimore","MD"),("Milwaukee","WI"),("Albuquerque","NM"),("Tucson","AZ"),
    ("Fresno","CA"),("Sacramento","CA"),
]
STREETS = ["Main St","Oak Ave","Medical Blvd","Health Dr","Park Rd","Elm St","Maple Ave"]

hcp_rows = []
hcp_id_counter = 1
for tid, tname, *_ in TERRITORIES:
    for bu, count in BU_HCP_QUOTA.items():
        specs = BU_SPECIALTIES[bu]
        for i in range(count):
            spec   = specs[i % len(specs)]
            fn     = random.choice(FIRST_NAMES)
            ln     = random.choice(LAST_NAMES)
            city, state = random.choice(CITIES_STATES)
            cred   = random.choices(CREDENTIALS, weights=CRED_WEIGHTS)[0]
            decile = random.randint(1, 10)
            seg    = "A" if decile >= 8 else ("B" if decile >= 5 else "C")
            hcp_rows.append({
                "CustomerID":   f"HCP{hcp_id_counter:04d}",
                "NPI":          f"1{random.randint(100000000, 999999999)}",
                "FirstName":    fn,
                "LastName":     ln,
                "DisplayName":  f"{cred}. {fn} {ln}",
                "Credential":   cred,
                "Specialty":    spec,
                "AddressLine1": f"{random.randint(100,9999)} {random.choice(STREETS)}",
                "City":         city,
                "State":        state,
                "ZipCode":      f"{random.randint(10000,99999)}",
                "CustomerType": "HCP",
                "TerritoryID":  tid,
                "DecileRank":   decile,
                "Segment":      seg,
                "ActiveFlag":   "Y",
            })
            hcp_id_counter += 1

HCO_NAMES = [
    "City Medical Center","Regional Health System","University Hospital",
    "Community Clinic Network","Metro Healthcare Group","Wellness Medical Associates",
]
for i, name in enumerate(HCO_NAMES, 1):
    city, state = random.choice(CITIES_STATES)
    tid = TERRITORIES[i % len(TERRITORIES)][0]
    hcp_rows.append({
        "CustomerID":   f"HCO{i:04d}",
        "NPI":          f"2{random.randint(100000000,999999999)}",
        "FirstName":    "",
        "LastName":     name,
        "DisplayName":  name,
        "Credential":   "",
        "Specialty":    "MULTI-SPECIALTY",
        "AddressLine1": f"{random.randint(100,9999)} Hospital Way",
        "City":         city,
        "State":        state,
        "ZipCode":      f"{random.randint(10000,99999)}",
        "CustomerType": "HCO",
        "TerritoryID":  tid,
        "DecileRank":   random.randint(7, 10),
        "Segment":      "A",
        "ActiveFlag":   "Y",
    })

df_master_customer = pd.DataFrame(hcp_rows)
hcp_ids = df_master_customer.loc[df_master_customer["CustomerType"] == "HCP", "CustomerID"].tolist()

# Lookup maps
hcp_spec_map  = dict(zip(df_master_customer["CustomerID"], df_master_customer["Specialty"]))
hcp_terr_map  = dict(zip(df_master_customer["CustomerID"], df_master_customer["TerritoryID"]))
hcp_decile_map= dict(zip(df_master_customer["CustomerID"], df_master_customer["DecileRank"]))

# ──────────────────────────────────────────────
# 4. DIM_SALES_REP
# ──────────────────────────────────────────────
rep_rows = []
for i, row in df_territory.iterrows():
    fn, ln = random.choice(FIRST_NAMES), random.choice(LAST_NAMES)
    rep_rows.append({
        "RepID":         f"REP{i+1:03d}",
        "RepFirstName":  fn,
        "RepLastName":   ln,
        "RepFullName":   f"{fn} {ln}",
        "TerritoryID":   row["TerritoryID"],
        "TerritoryName": row["TerritoryName"],
        "DistrictID":    row["DistrictID"],
        "DistrictName":  row["DistrictName"],
        "RegionID":      row["RegionID"],
        "RegionName":    row["RegionName"],
        "AreaID":        row["AreaID"],
        "AreaName":      row["AreaName"],
        "Role":          "Field Sales Representative",
        "Email":         f"{fn.lower()}.{ln.lower()}@pharmacore.com",
        "ManagerID":     f"MGR{(i//3)+1:03d}",
        "ActiveFlag":    "Y",
    })
df_sales_rep = pd.DataFrame(rep_rows)
rep_terr_map = dict(zip(df_sales_rep["RepID"], df_sales_rep["TerritoryID"]))

# ──────────────────────────────────────────────
# 5. DIM_SPEAKER  (top-decile HCPs as KOLs, BU-appropriate products)
# ──────────────────────────────────────────────
top_hcps = df_master_customer[
    (df_master_customer["CustomerType"] == "HCP") &
    (df_master_customer["DecileRank"] >= 8)
].head(12)

BU_SPEAKER_PRODS = {
    "AIR":      ["NEXOLID", "NEXOLID,VERITONEX", "VERITONEX"],
    "SBU":      ["CLAROZEPT", "CLAROZEPT,DEPTRAZOL", "DEPTRAZOL"],
    "Vaccines": ["NEXOLID"],  # cross-BU fallback — rare
}

spk_rows = []
for i, row in enumerate(top_hcps.itertuples(), 1):
    prog_dates = sorted([
        datetime(2025, random.randint(1, 6), random.randint(1, 28)).strftime("%Y-%m-%d")
        for _ in range(random.randint(2, 6))
    ])
    bu = SPEC_TO_BU.get(row.Specialty, "AIR")
    cert_opts = BU_SPEAKER_PRODS.get(bu, ["NEXOLID"])
    spk_rows.append({
        "SpeakerID":          f"SPK{i:03d}",
        "CustomerID":         row.CustomerID,
        "SpeakerName":        row.DisplayName,
        "Specialty":          row.Specialty,
        "TerritoryID":        row.TerritoryID,
        "SpeakerTier":        random.choice(["Tier 1", "Tier 1", "Tier 2"]),
        "CertifiedProducts":  random.choice(cert_opts),
        "Programs_YTD":       len(prog_dates),
        "LastProgramDate":    prog_dates[-1],
        "Attendees_Avg":      random.randint(8, 25),
        "Honorarium_YTD_USD": random.randint(3000, 25000),
        "ActiveFlag":         "Y",
    })
df_speaker = pd.DataFrame(spk_rows)

# ──────────────────────────────────────────────
# 6. FACT_CALLS
# ──────────────────────────────────────────────
CALL_TYPES    = ["Detail","Sample Drop","In-Service","Virtual Detail","MSL Meeting"]
CALL_W        = [50, 20, 10, 15, 5]
CALL_OUTCOMES = ["Positive","Neutral","Follow-up Required","Left Message","No See"]
OUTCOME_W     = [40, 25, 20, 10, 5]
CALL_NOTES    = ["Discussed efficacy data","Left samples","Reviewed clinical study",
                 "Patient case discussion","Formulary update","Follow up on AE",
                 "Presented new MOA data","Key account check-in","Peer insight shared"]

# HCPs grouped by territory
terr_hcp_map = {
    tid: df_master_customer[
        (df_master_customer["TerritoryID"] == tid) &
        (df_master_customer["CustomerType"] == "HCP")
    ]["CustomerID"].tolist()
    for tid in df_territory["TerritoryID"]
}

call_rows = []
call_id   = 1
for month in range(1, 7):
    for rep_id in df_sales_rep["RepID"]:
        terr_id    = rep_terr_map[rep_id]
        local_hcps = terr_hcp_map.get(terr_id, hcp_ids[:5]) or hcp_ids[:5]
        for _ in range(random.randint(18, 28)):
            call_dt = datetime(2025, month, random.randint(1, 28))
            cust_id = random.choice(local_hcps)
            # Pick a product that matches the HCP's BU
            hcp_bu  = SPEC_TO_BU.get(hcp_spec_map.get(cust_id, "NEUROLOGY"), "AIR")
            bu_prods = BU_PROD_IDS.get(hcp_bu, ["P001"])
            prod_id  = random.choice(bu_prods) if bu_prods else "P001"
            call_rows.append({
                "CallID":          f"CALL{call_id:06d}",
                "CallDate":        call_dt.strftime("%Y-%m-%d"),
                "CallYear":        2025,
                "CallMonth":       month,
                "CallPeriod":      f"2025-{month:02d}",
                "RepID":           rep_id,
                "CustomerID":      cust_id,
                "TerritoryID":     terr_id,
                "ProductID":       prod_id,
                "ProductFamily":   df_products.loc[df_products["ProductID"] == prod_id, "ProductFamily"].values[0],
                "CallType":        random.choices(CALL_TYPES, weights=CALL_W)[0],
                "CallOutcome":     random.choices(CALL_OUTCOMES, weights=OUTCOME_W)[0],
                "SamplesLeft":     random.choice([0, 2, 4, 6, 8]) if prod_id in ["P001", "P003"] else 0,
                "CallDurationMin": random.randint(5, 30),
                "NextCallDate":    (call_dt + timedelta(days=random.randint(7, 30))).strftime("%Y-%m-%d"),
                "CallNotes":       random.choice(CALL_NOTES),
            })
            call_id += 1

df_calls = pd.DataFrame(call_rows)

# ──────────────────────────────────────────────
# 7. FACT_PRESCRIPTIONS (BU-aligned — HCPs only prescribe their BU's products)
# ──────────────────────────────────────────────
PAYER_TYPES = ["Commercial","Medicare","Medicaid","HMO","Cash","Managed Care"]

rx_rows = []
rx_id   = 1
for month in range(1, 7):
    period = f"2025-{month:02d}"
    for hcp_id in hcp_ids:
        spec    = hcp_spec_map[hcp_id]
        bu      = SPEC_TO_BU.get(spec, "AIR")
        decile  = hcp_decile_map[hcp_id]
        terr_id = hcp_terr_map[hcp_id]
        prod_ids = BU_PROD_IDS.get(bu, [])

        if not prod_ids:
            continue  # Vaccines HCPs don't write pharma prescriptions

        base = decile * random.randint(5, 20)
        for prod_id in prod_ids:
            pfam    = df_products.loc[df_products["ProductID"] == prod_id, "ProductFamily"].values[0]
            trx     = max(0, int(base * random.uniform(0.5, 1.5)))
            nrx     = max(0, int(trx * random.uniform(0.10, 0.30)))
            trx_p   = max(0, int(trx * random.uniform(0.85, 1.15)))
            nrx_p   = max(0, int(nrx * random.uniform(0.85, 1.15)))
            mkt_trx = max(1, trx * random.randint(8, 15))
            mkt_nrx = max(1, nrx * random.randint(8, 15))
            rx_rows.append({
                "RxID":            f"RX{rx_id:08d}",
                "PeriodDate":      period,
                "PeriodYear":      2025,
                "PeriodMonth":     month,
                "CustomerID":      hcp_id,
                "ProductID":       prod_id,
                "ProductFamily":   pfam,
                "TerritoryID":     terr_id,
                "TRx":             trx,
                "NRx":             nrx,
                "TRx_Prior":       trx_p,
                "NRx_Prior":       nrx_p,
                "TRx_Change":      trx - trx_p,
                "NRx_Change":      nrx - nrx_p,
                "TRx_Change_Pct":  round((trx - trx_p) / trx_p * 100, 1) if trx_p else 0,
                "NRx_Change_Pct":  round((nrx - nrx_p) / nrx_p * 100, 1) if nrx_p else 0,
                "Market_TRx":      mkt_trx,
                "Market_NRx":      mkt_nrx,
                "TRx_Share_Pct":   round(trx / mkt_trx * 100, 2),
                "NRx_Share_Pct":   round(nrx / mkt_nrx * 100, 2),
                "PayerType":       random.choice(PAYER_TYPES),
                "PatientCount":    max(1, int(trx * random.uniform(0.65, 0.90))),
                "NewPatientCount": max(0, int(nrx * random.uniform(0.60, 0.90))),
            })
            rx_id += 1

df_prescriptions = pd.DataFrame(rx_rows)

# ──────────────────────────────────────────────
# 8. FACT_SALES_TARGET
# ──────────────────────────────────────────────
tgt_rows = []
tgt_id   = 1
PROD_RX_IDS = ["P001", "P003", "P005", "P007"]
for rep_id in df_sales_rep["RepID"]:
    terr_id = rep_terr_map[rep_id]
    for month in range(1, 7):
        period = f"2025-{month:02d}"
        for prod_id in PROD_RX_IDS:
            pfam       = df_products.loc[df_products["ProductID"] == prod_id, "ProductFamily"].values[0]
            actual_trx = int(df_prescriptions[
                (df_prescriptions["PeriodDate"] == period) &
                (df_prescriptions["TerritoryID"] == terr_id) &
                (df_prescriptions["ProductID"] == prod_id)
            ]["TRx"].sum())
            actual_nrx = int(df_prescriptions[
                (df_prescriptions["PeriodDate"] == period) &
                (df_prescriptions["TerritoryID"] == terr_id) &
                (df_prescriptions["ProductID"] == prod_id)
            ]["NRx"].sum())
            tgt_trx  = max(10, int(actual_trx * random.uniform(0.88, 1.18)))
            tgt_nrx  = max(2,  int(actual_nrx * random.uniform(0.88, 1.18)))
            unit_val = random.randint(50, 150)
            tgt_rows.append({
                "TargetID":             f"TGT{tgt_id:06d}",
                "PeriodDate":           period,
                "PeriodYear":           2025,
                "PeriodMonth":          month,
                "RepID":                rep_id,
                "TerritoryID":          terr_id,
                "ProductID":            prod_id,
                "ProductFamily":        pfam,
                "TRx_Target":           tgt_trx,
                "NRx_Target":           tgt_nrx,
                "TRx_Actual":           actual_trx,
                "NRx_Actual":           actual_nrx,
                "TRx_Attainment_Pct":   round(actual_trx / tgt_trx * 100, 1) if tgt_trx else 0,
                "NRx_Attainment_Pct":   round(actual_nrx / tgt_nrx * 100, 1) if tgt_nrx else 0,
                "SalesTarget_USD":      tgt_trx * unit_val,
                "SalesActual_USD":      actual_trx * unit_val,
                "SalesAttainment_Pct":  round(actual_trx / tgt_trx * 100, 1) if tgt_trx else 0,
            })
            tgt_id += 1

df_targets = pd.DataFrame(tgt_rows)

# ──────────────────────────────────────────────
# 9. KPI_TERRITORY
# ──────────────────────────────────────────────
kpi_terr_rows = []
for _, trow in df_territory.iterrows():
    tid  = trow["TerritoryID"]
    for month in range(1, 7):
        period = f"2025-{month:02d}"
        rx_m   = df_prescriptions[(df_prescriptions["TerritoryID"] == tid) & (df_prescriptions["PeriodDate"] == period)]
        call_m = df_calls[(df_calls["TerritoryID"] == tid) & (df_calls["CallPeriod"] == period)]
        tgt_m  = df_targets[(df_targets["TerritoryID"] == tid) & (df_targets["PeriodDate"] == period)]
        hcp_cnt= df_master_customer[(df_master_customer["TerritoryID"] == tid) & (df_master_customer["CustomerType"] == "HCP")]["CustomerID"].nunique()
        called = call_m["CustomerID"].nunique()

        kpi_terr_rows.append({
            "TerritoryID":            tid,
            "TerritoryName":          trow["TerritoryName"],
            "DistrictName":           trow["DistrictName"],
            "RegionName":             trow["RegionName"],
            "AreaName":               trow["AreaName"],
            "PeriodDate":             period,
            "PeriodMonth":            month,
            "CM_TRx":                 int(rx_m["TRx"].sum()),
            "CM_NRx":                 int(rx_m["NRx"].sum()),
            "CM_TRx_Prior":           int(rx_m["TRx_Prior"].sum()),
            "CM_NRx_Prior":           int(rx_m["NRx_Prior"].sum()),
            "CM_TRx_Change_Pct":      round((rx_m["TRx"].sum() - rx_m["TRx_Prior"].sum()) / max(1, rx_m["TRx_Prior"].sum()) * 100, 1),
            "CM_TRx_Share_Pct":       round(float(rx_m["TRx_Share_Pct"].mean()), 2) if len(rx_m) else 0,
            "CM_NRx_Share_Pct":       round(float(rx_m["NRx_Share_Pct"].mean()), 2) if len(rx_m) else 0,
            "CM_Market_TRx":          int(rx_m["Market_TRx"].sum()),
            "CM_Calls":               len(call_m),
            "CM_TRx_Attainment_Pct":  round(float(tgt_m["TRx_Attainment_Pct"].mean()), 1) if len(tgt_m) else 0,
            "CM_Sales_Attainment_Pct":round(float(tgt_m["SalesAttainment_Pct"].mean()), 1) if len(tgt_m) else 0,
            "CM_NewPatients":         int(rx_m["NewPatientCount"].sum()),
            "HCP_Universe":           hcp_cnt,
            "HCP_Called_Pct":         round(called / max(1, hcp_cnt) * 100, 1),
            "Call_Frequency_Avg":     round(len(call_m) / max(1, called), 2),
            # Per-BU product TRx (key columns for BU filtering)
            "NEXOLID_TRx":            int(rx_m[rx_m["ProductFamily"] == "NEXOLID"]["TRx"].sum()),
            "VERITONEX_TRx":          int(rx_m[rx_m["ProductFamily"] == "VERITONEX"]["TRx"].sum()),
            "CLAROZEPT_TRx":          int(rx_m[rx_m["ProductFamily"] == "CLAROZEPT"]["TRx"].sum()),
            "DEPTRAZOL_TRx":          int(rx_m[rx_m["ProductFamily"] == "DEPTRAZOL"]["TRx"].sum()),
            "NEXOLID_NRx":            int(rx_m[rx_m["ProductFamily"] == "NEXOLID"]["NRx"].sum()),
            "VERITONEX_NRx":          int(rx_m[rx_m["ProductFamily"] == "VERITONEX"]["NRx"].sum()),
            "CLAROZEPT_NRx":          int(rx_m[rx_m["ProductFamily"] == "CLAROZEPT"]["NRx"].sum()),
            "DEPTRAZOL_NRx":          int(rx_m[rx_m["ProductFamily"] == "DEPTRAZOL"]["NRx"].sum()),
        })

df_kpi_territory = pd.DataFrame(kpi_terr_rows)

# ──────────────────────────────────────────────
# 10. KPI_HCP360
# ──────────────────────────────────────────────
hcp360_rows = []
LATEST = "2025-06"
for hcp_id in hcp_ids:
    info     = df_master_customer.loc[df_master_customer["CustomerID"] == hcp_id].iloc[0]
    rx_cm    = df_prescriptions[(df_prescriptions["CustomerID"] == hcp_id) & (df_prescriptions["PeriodDate"] == LATEST)]
    rx_ytd   = df_prescriptions[df_prescriptions["CustomerID"] == hcp_id]
    calls_cm = df_calls[(df_calls["CustomerID"] == hcp_id) & (df_calls["CallPeriod"] == LATEST)]
    calls_ytd= df_calls[df_calls["CustomerID"] == hcp_id]
    last_call= calls_ytd["CallDate"].max() if len(calls_ytd) else "No Call"

    def pfam_rx(data, fam, col):
        v = data.loc[data["ProductFamily"] == fam, col].sum()
        return int(v) if v else 0

    def pfam_share(data, fam):
        v = data.loc[data["ProductFamily"] == fam, "TRx_Share_Pct"].mean()
        return round(float(v), 2) if not (isinstance(v, float) and np.isnan(v)) else 0.0

    hcp360_rows.append({
        "CustomerID":             hcp_id,
        "NPI":                    info["NPI"],
        "DisplayName":            info["DisplayName"],
        "Specialty":              info["Specialty"],
        "Credential":             info["Credential"],
        "City":                   info["City"],
        "State":                  info["State"],
        "TerritoryID":            info["TerritoryID"],
        "DecileRank":             info["DecileRank"],
        "Segment":                info["Segment"],
        # Current Month Rx totals (only BU products for this HCP)
        "CM_Total_TRx":           int(rx_cm["TRx"].sum()),
        "CM_Total_NRx":           int(rx_cm["NRx"].sum()),
        "CM_Total_Patients":      int(rx_cm["PatientCount"].sum()),
        "CM_New_Patients":        int(rx_cm["NewPatientCount"].sum()),
        "CM_TRx_Prior":           int(rx_cm["TRx_Prior"].sum()),
        "CM_TRx_Change_Pct":      round((rx_cm["TRx"].sum() - rx_cm["TRx_Prior"].sum()) / max(1, rx_cm["TRx_Prior"].sum()) * 100, 1),
        # Per product family (AIR HCPs have NEXOLID/VERITONEX > 0; SBU HCPs have CLAROZEPT/DEPTRAZOL > 0)
        "NEXOLID_TRx_CM":         pfam_rx(rx_cm, "NEXOLID",   "TRx"),
        "NEXOLID_NRx_CM":         pfam_rx(rx_cm, "NEXOLID",   "NRx"),
        "NEXOLID_TRxShare_Pct":   pfam_share(rx_cm, "NEXOLID"),
        "VERITONEX_TRx_CM":       pfam_rx(rx_cm, "VERITONEX", "TRx"),
        "VERITONEX_NRx_CM":       pfam_rx(rx_cm, "VERITONEX", "NRx"),
        "VERITONEX_TRxShare_Pct": pfam_share(rx_cm, "VERITONEX"),
        "CLAROZEPT_TRx_CM":       pfam_rx(rx_cm, "CLAROZEPT", "TRx"),
        "CLAROZEPT_NRx_CM":       pfam_rx(rx_cm, "CLAROZEPT", "NRx"),
        "CLAROZEPT_TRxShare_Pct": pfam_share(rx_cm, "CLAROZEPT"),
        "DEPTRAZOL_TRx_CM":       pfam_rx(rx_cm, "DEPTRAZOL", "TRx"),
        "DEPTRAZOL_NRx_CM":       pfam_rx(rx_cm, "DEPTRAZOL", "NRx"),
        "DEPTRAZOL_TRxShare_Pct": pfam_share(rx_cm, "DEPTRAZOL"),
        # YTD
        "YTD_TRx":                int(rx_ytd["TRx"].sum()),
        "YTD_NRx":                int(rx_ytd["NRx"].sum()),
        "YTD_TRx_Share_Avg":      round(float(rx_ytd["TRx_Share_Pct"].mean()), 2) if len(rx_ytd) else 0.0,
        # Call Activity
        "CM_Calls":               len(calls_cm),
        "YTD_Calls":              len(calls_ytd),
        "Last_Call_Date":         last_call,
        "CM_Call_Outcome_Last":   calls_cm["CallOutcome"].iloc[-1] if len(calls_cm) else "No Call",
        "Is_Speaker":             "Y" if hcp_id in df_speaker["CustomerID"].values else "N",
    })

df_hcp360 = pd.DataFrame(hcp360_rows)

# ──────────────────────────────────────────────
# Excel Writer Helpers
# ──────────────────────────────────────────────
HDR_COLORS = {
    "dim":  "1F4E79",
    "fact": "833C00",
    "kpi":  "375623",
}

def write_sheet(ws, data, hdr_color="1F4E79"):
    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    fill = PatternFill(start_color=hdr_color, end_color=hdr_color, fill_type="solid")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)
    ws.freeze_panes = "A2"

# ──────────────────────────────────────────────
# Build workbooks
# ──────────────────────────────────────────────
SHEETS = [
    ("Dim_Territory",      df_territory,       "dim"),
    ("Dim_Products",       df_products,        "dim"),
    ("Dim_MasterCustomer", df_master_customer, "dim"),
    ("Dim_SalesRep",       df_sales_rep,       "dim"),
    ("Dim_Speaker",        df_speaker,         "dim"),
    ("Fact_Calls",         df_calls,           "fact"),
    ("Fact_Prescriptions", df_prescriptions,   "fact"),
    ("Fact_SalesTarget",   df_targets,         "fact"),
    ("KPI_Territory",      df_kpi_territory,   "kpi"),
    ("KPI_HCP360",         df_hcp360,          "kpi"),
]

wb = Workbook()
wb.remove(wb.active)
for name, data, kind in SHEETS:
    ws = wb.create_sheet(title=name)
    write_sheet(ws, data, HDR_COLORS[kind])
combined_path = "sourcedata/IQ2020_POC_DataModel.xlsx"
wb.save(combined_path)
print(f"Saved: {combined_path}")

for name, data, kind in SHEETS:
    wb_s = Workbook()
    ws_s = wb_s.active
    ws_s.title = name
    write_sheet(ws_s, data, HDR_COLORS[kind])
    path = f"sourcedata/{name}.xlsx"
    wb_s.save(path)
    print(f"Saved: {path}")

# ──────────────────────────────────────────────
# Summary
# ──────────────────────────────────────────────
print("\n" + "=" * 60)
print(" IQ2020 POC -- DATA MODEL SUMMARY")
print("=" * 60)
print(f"{'Sheet':<26} {'Rows':>6}  {'Cols':>5}  {'Type'}")
print("-" * 60)
for name, data, kind in SHEETS:
    print(f"{name:<26} {len(data):>6}  {len(data.columns):>5}  {kind.upper()}")
print("=" * 60)
print(f"\nHCPs per territory : {sum(BU_HCP_QUOTA.values())} ({dict(BU_HCP_QUOTA)})")
print(f"Total HCPs         : {len(hcp_ids)}")
print(f"Total Calls        : {len(df_calls):,}")
print(f"Total Rx records   : {len(df_prescriptions):,}")
print(f"Total Targets      : {len(df_targets):,}")
print(f"Territories        : {len(df_territory)}")
print(f"Date range         : Jan 2025 - Jun 2025 (6 months)")
print(f"\nBU Specialty mapping (exclusive):")
for bu, specs in BU_SPECIALTIES.items():
    prods = [df_products.loc[df_products['ProductID']==p,'ProductFamily'].values[0] for p in BU_PROD_IDS[bu]]
    print(f"  {bu:<10} specialties={specs}  products={prods}")
print(f"\nAll files saved to: sourcedata/")
